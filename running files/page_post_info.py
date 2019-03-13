import os,sys
import json
import facebook
import xlsxwriter
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd

if __name__ == '__main__':
    token = "EAANXc609TdkBAO3HmSoswBZCTIbmZBMOcdzvLa8c97fdDZBzCjZCL2vAhJYPhyKt5sURY5VlozyHOZABZB6lxrPU5Bb8jM0PLFHh0xCj376nqu6EQZA6PoGbnI1cKyGYiOtrNNyLUebm55GGjNGI5VL6Tj1R9IstsIUSQHBbW7WVP7ZBUbZAn4occ"
    
    graph = facebook.GraphAPI(access_token=token, version = 3.0)
    ##Page Post
    page_posts_impressions = graph.get_object(id='974146599436745', fields='insights.metric(page_posts_impressions).period(week)')

    page_post_engagements = graph.get_object(id='974146599436745', fields='insights.metric(page_post_engagements).period(week)')

    page_consumptions = graph.get_object(id='974146599436745', fields='insights.metric(page_consumptions).period(week)')

    page_posts_impressions_paid = graph.get_object(id='974146599436745', fields='insights.metric(page_posts_impressions_paid).period(week)')

    page_posts_impressions_organic = graph.get_object(id='974146599436745', fields='insights.metric(page_posts_impressions_organic).period(week)')

    page_posts_impressions_viral = graph.get_object(id='974146599436745', fields='insights.metric(page_posts_impressions_viral).period(week)')
  
def put_zeros_in_empty_cells(file_name):
            # load demo.xlsx 
        wb=load_workbook(file_name)
            # select demo.xlsx
        sheet=wb.active
            # get max row count
        max_row=sheet.max_row
            # get max column count
        max_column=sheet.max_column
        for i in range(1,max_row+1):
        # iterate over all columns
            for j in range(1,max_column+1):
                # get particular cell value    
                if sheet.cell(row=i,column=j).value==None:
                    sheet.cell(row=i,column=j).value = 0
        wb.save(file_name)

def create_page_post_info():
    
    def get_page_post_all():
        date = page_posts_impressions["insights"]["data"][0]['values'][1]['end_time']
        
        dictionary_title_values = {}
        page_posts_impressions_value = page_posts_impressions["insights"]["data"][0]['values'][1]['value']
        dictionary_title_values.update({"Page Post Impressions":page_posts_impressions_value})
        
        page_post_engagements_value = page_post_engagements["insights"]["data"][0]['values'][1]['value']
        dictionary_title_values.update({"Page Post Engagements":page_post_engagements_value})
        
        page_consumptions_value = page_consumptions["insights"]["data"][0]['values'][1]['value']
        dictionary_title_values.update({"Page Consumptios":page_consumptions_value})
        
        page_posts_impressions_paid_value = page_posts_impressions_paid["insights"]["data"][0]['values'][1]['value']
        dictionary_title_values.update({"Page Post Impressions Paid":page_posts_impressions_paid_value})
        
        page_posts_impressions_organic_value = page_posts_impressions_organic["insights"]["data"][0]['values'][1]['value']
        dictionary_title_values.update({"Page Post Impressions Organic":page_posts_impressions_organic_value})
        
        page_posts_impressions_viral_value = page_posts_impressions_viral["insights"]["data"][0]['values'][1]['value']
        dictionary_title_values.update({"Page Post Impressions Viral":page_posts_impressions_viral_value})
        
        page_post_path_direc="excels/lite/"+"lite-Page-Post.xlsx"
        
        if os.path.isfile(page_post_path_direc):
            update_page_post_info_in_xlsx(date,dictionary_title_values,page_post_path_direc)
        else:
            write_page_post_info_in_xlsx(date,dictionary_title_values,page_post_path_direc)
  
    def write_page_post_info_in_xlsx(date,dictionary,name_xlsx):
        workbook = xlsxwriter.Workbook(name_xlsx)
        worksheet = workbook.add_worksheet()
        #writting first line with date 
        worksheet.write(0, 0,'Date-1 week until')
        date_to_write= date.split("T",1)[0] #old
        date_to_write = datetime.strptime(date_to_write, '%Y-%m-%d').strftime("%-m/%-d/%Y")
        worksheet.write(1, 0,date_to_write)
        All_titles = list(dictionary.keys())
        row=0
        col=1
        #writting first and second line with gender-age indicators & values of ages-gender
        for i in range(len(All_titles)):
            
            worksheet.write(row,col,All_titles[i])
            worksheet.write(row+1,col,dictionary[All_titles[i]])
            col +=1
        workbook.close()
        print(name_xlsx + " is created!")

    def update_page_post_info_in_xlsx(date,dictionary,name_xlsx):
        wb2 = load_workbook(name_xlsx)
        ws = wb2.active
        last_row = ws.max_row
        date_to_write= date.split("T",1)[0] 
        date_to_write = datetime.strptime(date_to_write, '%Y-%m-%d').strftime("%-d/%-m/%Y")
        ws.cell(last_row+1,1).value = date_to_write

        All_titles = list(dictionary.keys())
        col=2
        for i in range(len(All_titles)):
            ws.cell(last_row+1,col).value = dictionary[All_titles[i]]
            col +=1

        wb2.save(name_xlsx)
        #put_zeros_in_empty_cells(name_xlsx)
        put_zeros_in_empty_cells(name_xlsx)
        print(name_xlsx + " is updated!")

    get_page_post_all()

create_page_post_info()