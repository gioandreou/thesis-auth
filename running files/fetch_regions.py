from openpyxl import load_workbook
import numpy as np
import os,sys
import xlsxwriter
import datetime as dt
import matplotlib.pyplot as plt
import pandas as pd
from datetime import datetime
from tabulate import tabulate
from pandas import ExcelWriter

def striplist(l):
    return([x.strip() for x in l])

def fetch_xlsx(dataframe):
    #DROP THE DATE COLUMN TO KEEP ONLY THE CITY-REGION VALUES
    dataframe = dataframe.drop(['Date'], axis=1)
    headers=list(dataframe.columns.values)
    
    #KEEP THE SECOND PART OF THE STRING e.g. "Thessaloniki, Central Macedonia, Greece" -> " Central Macedonia, Greece"
    headers = [item.replace(item, item.split(",",1)[1]) for item in headers]
    
    #DELETE THE LEADING WHITESPACE FROM THE STRING e.g. " Central Macedonia, Greece"-> "Central Macedonia, Greece"
    headers = striplist(headers)
    
    with pd.option_context('display.max_rows', None, 'display.max_columns', None):
        dataframe.columns = headers
        #GROUP DATA BY THEIR COLUMNS(REGIONS) AND SUM TOGETHER THE COLUMNS(REGIONS) WITH THE SAME NAME .
        dataframe = dataframe.groupby(dataframe.columns, axis=1).sum()
       
        #CALCULATE THE MEAN VALUE FOR EVERY REGION
        dataframe_mean = dataframe.mean(axis=0 )
        #KEEP THE N LARGEST REGIONS 
        dataframe_mean = dataframe_mean.nlargest(7)    
        
        #COMMENTED OUT OPTION, TO KEEP THE LATEST VALUES OF THE DATAFRAME
        
        #Latest values of regions
        dataframe_last = dataframe.drop(dataframe.index[:-1])
        dataframe_last = dataframe_last.transpose()
        column_name=list(dataframe_last.columns.values)
        dataframe_last = dataframe_last.nlargest(7,column_name[0])    
        
        if os.path.isfile('excels/lite/RegionDF_countinuously.xlsx'):
                update_regions_continuously(dataframe_last)
        else:
                create_regions_continuously(dataframe_last)

        create_regions_simple(dataframe_mean)
         

def create_regions_continuously(dataframe):
        name_xlsx = 'excels/lite/RegionDF_countinuously.xlsx'
        workbook = xlsxwriter.Workbook(name_xlsx)
        worksheet = workbook.add_worksheet()
        worksheet.write(0, 0,'Date Fetched')
        worksheet.write(0, 1,'Region')
        worksheet.write(0, 2,'Fans')

        todays_date = datetime.today().strftime("%d/%m/%Y")

        region_list = dataframe.index.tolist()
        value_list = dataframe[dataframe.columns[0]].tolist()
        row=1

        for i in range(len(region_list)):
                worksheet.write(row+i,0,todays_date) #write date of fetching 
                worksheet.write(row+i,1,region_list[i]) #write each region in xlsx 2nd col each row
                worksheet.write(row+i,2,value_list[i])  #write each region's value in 3rd col of xlsx

        workbook.close()
        print(name_xlsx + " is created!")

def update_regions_continuously(dataframe):
        name_xlsx = 'excels/lite/RegionDF_countinuously.xlsx'
        workbook = load_workbook(name_xlsx)
        worksheet = workbook.active
        max_row = worksheet.max_row
        max_row=max_row+1

        todays_date = datetime.today().strftime("%d/%m/%Y")

        region_list = dataframe.index.tolist()
        value_list = dataframe[dataframe.columns[0]].tolist()

        for i in range(len(region_list)):

                worksheet.cell(row=max_row+i,column=1).value=todays_date
                worksheet.cell(row=max_row+i,column=2).value=region_list[i]
                worksheet.cell(row=max_row+i,column=3).value=value_list[i]
        workbook.save(name_xlsx)
        print(name_xlsx + " is updated!")

def create_regions_simple(dataframe):
        #WRITE THE DF TO XLSX IN ORDER TO BE PARSED BY OTHER FILES 
        name_xlsx = 'excels/lite/RegionDF.xlsx'
        writer = ExcelWriter(name_xlsx)
        dataframe.to_excel(writer,'Sheet1')
        writer.save()        
        
        wb2 = load_workbook(name_xlsx)
        ws = wb2.active
        ws.cell(row=1,column=1).value='Regions'
        ws.cell(row=1,column=2).value='Fans'
        wb2.save(name_xlsx)
        print(name_xlsx + " is created!")

def main():

        #PARSE THE XLSX TO DATAFRAME
    city_region = pd.read_excel('excels/lite/lite-City.xlsx')
    fetch_xlsx(city_region)

main()