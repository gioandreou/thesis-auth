import os,sys
import json
import facebook
import xlsxwriter
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd
from unidecode import unidecode


if __name__ == '__main__':
    token = "EAANXc609TdkBAO3HmSoswBZCTIbmZBMOcdzvLa8c97fdDZBzCjZCL2vAhJYPhyKt5sURY5VlozyHOZABZB6lxrPU5Bb8jM0PLFHh0xCj376nqu6EQZA6PoGbnI1cKyGYiOtrNNyLUebm55GGjNGI5VL6Tj1R9IstsIUSQHBbW7WVP7ZBUbZAn4occ"
    
    graph = facebook.GraphAPI(access_token=token, version = 3.0)

    posts = graph.get_object(id='974146599436745', fields='posts')


def get_post_information():
    list_post = posts["posts"]["data"]
    #print (json.dumps(list_post, indent=4))
    dict_of_posts = {}
    for item in reversed(list_post):
        if "message" in item:
            id_of_post= item["id"]
            date_to_write= item["created_time"].split("+",1)[0]
            #date_to_write = datetime.strptime(date_to_write, '%Y-%m-%d').strftime("%d/%m/%Y")
            ###All People
            ##Post Impressions = The number of times your Page's post entered a person's screen.
            post_impressions = graph.get_object(id=id_of_post, fields='insights.metric(post_impressions)')
            post_impressions_value = post_impressions["insights"]["data"][0]['values'][0]['value']
            
            #Post Impressions paid unique = The number of people who had your Page's post enter their screen through paid distribution such as an ad
            post_impressions_paid_unique = graph.get_object(id=id_of_post, fields='insights.metric(post_impressions_paid_unique)')
            post_impressions_paid_unique_value = post_impressions_paid_unique["insights"]["data"][0]['values'][0]['value']
            
            #Post Impressions Organic Unique = The number of times your Page's posts entered a person's screen through unpaid distribution
            post_impressions_organic_unique = graph.get_object(id=id_of_post, fields='insights.metric(post_impressions_organic_unique)')
            post_impressions_organic_unique_value = post_impressions_organic_unique["insights"]["data"][0]['values'][0]['value']
            
            ###Fans
            #Post Impressions Fans = The number of people who have like your Page who saw your Page post.
            post_impressions_fan_unique = graph.get_object(id=id_of_post, fields='insights.metric(post_impressions_fan_unique)')
            post_impressions_fan_unique_value = post_impressions_fan_unique["insights"]["data"][0]['values'][0]['value']
            
            #Post Impressions Fan Paid = The number of impressions for your Page post by people who like your Page in an Ad or Sponsored Story.
            post_impressions_fan_paid = graph.get_object(id=id_of_post, fields='insights.metric(post_impressions_fan_paid)')
            post_impressions_fan_paid_value = post_impressions_fan_paid["insights"]["data"][0]['values'][0]['value']

            #Post Enganged Users = The number of people who clicked anywhere in your posts.
            post_engaged_users = graph.get_object(id=id_of_post, fields='insights.metric(post_engaged_users)')
            post_engaged_users_value = post_engaged_users["insights"]["data"][0]['values'][0]['value']
            
            #Post Impressions Viral = The number of times your Page's post entered a person's screen with social information attached. 
            post_impressions_viral_unique = graph.get_object(id=id_of_post, fields='insights.metric(post_impressions_viral_unique)')
            post_impressions_viral_unique_value = post_impressions_viral_unique["insights"]["data"][0]['values'][0]['value']


            ##Reactions
            post_reactions_by_type_total = graph.get_object(id=id_of_post, fields='insights.metric(post_reactions_by_type_total)')
            post_reactions_by_type_total_dict_of_values = post_reactions_by_type_total["insights"]["data"][0]['values'][0]['value']
            total_of_reactions = 0
            for key, value in post_reactions_by_type_total_dict_of_values.items():
                    total_of_reactions = total_of_reactions +value 

            temp_dict={id_of_post:[date_to_write,unidecode(item["message"]),
            post_impressions_value,
            post_impressions_paid_unique_value,
            post_impressions_organic_unique_value,
            post_impressions_fan_unique_value,
            post_impressions_fan_paid_value,
            post_engaged_users_value,
            post_impressions_viral_unique_value,
            total_of_reactions]}

            dict_of_posts.update(temp_dict)
                    
    post_info_path_direc="excels/lite/"+"lite-Post-Info-Countinuously.xlsx"
    #if file exists update the file with new dates and posts' infos or if it doesnt create a new one 
    if os.path.isfile(post_info_path_direc):
        update_post_info_in_xlsx(dict_of_posts,post_info_path_direc)
        update_status_in_xlsx(dict_of_posts,post_info_path_direc)
        
    else:
        write_post_info_in_xlsx(dict_of_posts,post_info_path_direc)    
    


def write_post_info_in_xlsx(dictionary,name_xlsx):
    workbook = xlsxwriter.Workbook(name_xlsx)
    worksheet = workbook.add_worksheet()
    #writting first line with ID, date and Title of values 
    worksheet.write(0, 0,'Date Fetched')
    worksheet.write(0, 1,'ID')
    worksheet.write(0, 2,'Date Created')
    worksheet.write(0, 3,'Message')
    worksheet.write(0, 4,'Impressions')
    worksheet.write(0, 5,'Impressions Paid')
    worksheet.write(0, 6,'Impressions Organic')
    worksheet.write(0, 7,'Impressions Fans')
    worksheet.write(0, 8,'Impressions Fans Paid')
    worksheet.write(0, 9,'Enganged Users')
    worksheet.write(0, 10,'Impressions Viral')
    worksheet.write(0, 11,'Total Reactions')
    worksheet.write(0, 12,'STATUS')


    All_ids = list(dictionary.keys())
    col=1
    row=1

    todays_date = datetime.today().strftime("%d/%m/%Y")
    
    for i in range(len(All_ids)): #every id-post 
        worksheet.write(row+i,0,todays_date) #write date of fetching 
        worksheet.write(row+i,1,All_ids[i]) #write id in xlsx 1st col each row
        
        for j in range(1,len(dictionary[All_ids[i]])+1): #all the infos of every post-id
            worksheet.write(row+i,col+j,dictionary[All_ids[i]][j-1]) #write every value from the list of values inside the dictionary
        worksheet.write(row+i,len(dictionary[All_ids[i]])+2,'START')
    workbook.close()
    print(name_xlsx + " is created!")


def update_status_in_xlsx(dictionary,name_xlsx):
    workbook = load_workbook(name_xlsx)
    worksheet = workbook.active

    max_row = worksheet.max_row

    all_ids_of_posts = list(dictionary.keys()) #all ids of posts that have been posted since latest fetch
    dict_id_position = {} # dictionary with key = id and value = [positions(rows) that this id is appeared in xlsx]

    for item_id in all_ids_of_posts: #iterate through all id items 
        id_positions=[]
        for row in range(2,max_row+1): # for every item, check every row if this id is appeared
            if item_id == worksheet.cell(row=row,column=2).value: # if yes append the row value at the list
                id_positions.append(row)
        
        dict_id_position.update({item_id:id_positions})

    print(dict_id_position)

    for key,value in dict_id_position.items():
        for position in range(len(value)-1) :
            if (worksheet.cell(row=value[position],column=5).value<worksheet.cell(row=value[position+1],column=5).value):
                worksheet.cell(row=value[position+1],column=13).value='ALIVE'
            else :
                worksheet.cell(row=value[position+1],column=13).value='DEAD'
    workbook.save(name_xlsx)      


def update_post_info_in_xlsx(dictionary,name_xlsx):
    wb2 = load_workbook(name_xlsx)
    ws = wb2.active
    last_row = ws.max_row
    
    All_ids = list(dictionary.keys())
    col=2
    row=last_row+1

    todays_date = datetime.today().strftime("%d/%m/%Y")
    
    for i in range(len(All_ids)): #every id-post 
        ws.cell(row=row+i,column=1).value=todays_date #write date of fetching 
        ws.cell(row=row+i,column=2).value=All_ids[i] #write id in xlsx 1st col each row
        
        for j in range(1,len(dictionary[All_ids[i]])+1): #all the infos of every post-id
            ws.cell(row=row+i,column=col+j).value=dictionary[All_ids[i]][j-1] #write every value from the list of values inside the dictionary
    wb2.save(name_xlsx)


    print(name_xlsx + " is updated!")


get_post_information()