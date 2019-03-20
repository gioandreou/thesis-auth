import os,sys
import json
import facebook
import xlsxwriter
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd

date_format = "%Y-%m-%d"

if __name__ == '__main__':
    token = "EAANXc609TdkBAO3HmSoswBZCTIbmZBMOcdzvLa8c97fdDZBzCjZCL2vAhJYPhyKt5sURY5VlozyHOZABZB6lxrPU5Bb8jM0PLFHh0xCj376nqu6EQZA6PoGbnI1cKyGYiOtrNNyLUebm55GGjNGI5VL6Tj1R9IstsIUSQHBbW7WVP7ZBUbZAn4occ"
    
    graph = facebook.GraphAPI(access_token=token, version = 3.0)
    ##Page info Everything is at week period. So each date is calculated from the date shown and 7 days back

    #page_fans_by_like_source = graph.get_object(id='974146599436745', fields='insights.metric(page_fans_by_like_source)')

    #The number of people who had any content from your Page or about your Page enter their screen. 
    page_impressions_unique = graph.get_object(id='974146599436745', fields='insights.metric(page_impressions_unique).period(week)')

    #The number of people who had any content from your Page or about your Page enter their screen through paid distribution such as an ad.
    page_impressions_paid_unique = graph.get_object(id='974146599436745', fields='insights.metric(page_impressions_paid_unique).period(week)')

    #The number of people who had any content from your Page or about your Page enter their screen through unpaid distribution.
    page_impressions_organic_unique = graph.get_object(id='974146599436745', fields='insights.metric(page_impressions_organic_unique).period(week)')

    #The number of new people who have liked your page broken down by paid and non-paid. only day
    page_fan_adds_by_paid_non_paid_unique = graph.get_object(id='974146599436745', fields='insights.metric(page_fan_adds_by_paid_non_paid_unique)')
    
    #The number of times a Page's profile has been viewed by logged in and logged out people.
    page_views_total= graph.get_object(id='974146599436745', fields='insights.metric(page_views_total).period(week)')

    #The number of users who like the Page
    fan_count = graph.get_object(id='974146599436745', fields='insights.metric(page_fans)')


def put_zeros_in_empty_cells(file_name):
            # load demo.xlsx 
        wb=load_workbook(file_name)
            # select demo.xlsx
        sheet=wb.active
            # get max row count
        max_row=sheet.max_row
            # get max column count
        max_column=sheet.max_column
        for i in range(1,max_row+1):
        # iterate over all columns
            for j in range(1,max_column+1):
                # get particular cell value    
                if sheet.cell(row=i,column=j).value==None:
                    sheet.cell(row=i,column=j).value = 0
        wb.save(file_name)

def create_page_info():
    
    def get_page_info():
            date = page_impressions_unique["insights"]["data"][0]['values'][1]['end_time']

            dictionary_title_values = {}
            '''
            page_fans_by_like_source_value = page_fans_by_like_source["insights"]["data"][0]['values'][1]['value']
            dictionary_title_values.update({"Page Fans By Like Source":page_fans_by_like_source_value})
            '''
            page_views_total_value = page_views_total["insights"]["data"][0]['values'][1]['value']
            dictionary_title_values.update({"Page View Totals":page_views_total_value})


            fan_count_value = fan_count["insights"]["data"][0]['values'][1]['value']
            dictionary_title_values.update({"Page Fans":fan_count_value})

            page_fan_adds_by_paid_value = page_fan_adds_by_paid_non_paid_unique["insights"]["data"][0]['values'][1]['value']['paid']
            dictionary_title_values.update({"Page Fan Adds Paid":page_fan_adds_by_paid_value})
            
            page_fan_adds_by_non_paid_value = page_fan_adds_by_paid_non_paid_unique["insights"]["data"][0]['values'][1]['value']['unpaid']
            dictionary_title_values.update({"Page Fan Adds Non Paid":page_fan_adds_by_non_paid_value})

            page_impressions_unique_value = page_impressions_unique["insights"]["data"][0]['values'][1]['value']
            dictionary_title_values.update({"Page Impressions ":page_impressions_unique_value})
            
            page_impressions_paid_unique_value = page_impressions_paid_unique["insights"]["data"][0]['values'][1]['value']
            dictionary_title_values.update({"Page Impressions Paid":page_impressions_paid_unique_value})
            
            page_impressions_organic_unique_value = page_impressions_organic_unique["insights"]["data"][0]['values'][1]['value']
            dictionary_title_values.update({"Page Impressions Organic":page_impressions_organic_unique_value})


            #print (json.dumps(dictionary_title_values, indent=4))
            page_info_path_direc="excels/lite/"+"lite-Page-Info.xlsx"
            
            if os.path.isfile(page_info_path_direc):
                update_page_post_info_in_xlsx(date,dictionary_title_values,page_info_path_direc)
            else:
                write_page_post_info_in_xlsx(date,dictionary_title_values,page_info_path_direc)
            
    def write_page_post_info_in_xlsx(date,dictionary,name_xlsx):
        workbook = xlsxwriter.Workbook(name_xlsx)
        worksheet = workbook.add_worksheet()
        #writting first line with date 
        worksheet.write(0, 0,'Date')
        date_to_write= date.split("T",1)[0] #old
        date_to_write = datetime.strptime(date_to_write, '%Y-%m-%d').strftime(date_format)
        worksheet.write(1, 0,date_to_write)
        All_titles = list(dictionary.keys())
        row=0
        col=1
        #writting first line with titles and second with their values
        for i in range(len(All_titles)):
            
            worksheet.write(row,col,All_titles[i])
            worksheet.write(row+1,col,dictionary[All_titles[i]])
            col +=1
        workbook.close()
        print(name_xlsx + " is created!")

    def update_page_post_info_in_xlsx(date,dictionary,name_xlsx):
        wb2 = load_workbook(name_xlsx)
        ws = wb2.active
        last_row = ws.max_row
        date_to_write= date.split("T",1)[0]
        date_to_write = datetime.strptime(date_to_write, '%Y-%m-%d').strftime(date_format)
        ws.cell(last_row+1,1).value = date_to_write

        All_titles = list(dictionary.keys())
        col=2
        for i in range(len(All_titles)):
            ws.cell(last_row+1,col).value = dictionary[All_titles[i]]
            col +=1

        wb2.save(name_xlsx)
        #put_zeros_in_empty_cells(name_xlsx)
        put_zeros_in_empty_cells(name_xlsx)
        print(name_xlsx + " is updated!")
    
    get_page_info()

create_page_info()
