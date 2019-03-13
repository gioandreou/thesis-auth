import os,sys
import json
import facebook
import xlsxwriter
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd

date_format = "%-d/%-m/%y"

if __name__ == '__main__':
    token = "EAANXc609TdkBAO3HmSoswBZCTIbmZBMOcdzvLa8c97fdDZBzCjZCL2vAhJYPhyKt5sURY5VlozyHOZABZB6lxrPU5Bb8jM0PLFHh0xCj376nqu6EQZA6PoGbnI1cKyGYiOtrNNyLUebm55GGjNGI5VL6Tj1R9IstsIUSQHBbW7WVP7ZBUbZAn4occ"
    
    graph = facebook.GraphAPI(access_token=token, version = 3.0)

    #page_fans_gender_locale = Aggregated language data about the people who like your Page based on the default language setting selected when accessing Facebook.
    page_fans_locale = graph.get_object(id='974146599436745', fields='insights.metric(page_fans_locale)')
    
    #page_fans_city = Aggregated Facebook location data, sorted by city, about the people who like your Page.
    page_fans_city = graph.get_object(id='974146599436745', fields='insights.metric(page_fans_city)')
    
    #page_fans_city = The number of people, aggregated per country, that like your Page.
    page_fans_country = graph.get_object(id='974146599436745', fields='insights.metric(page_fans_country)')
    
    #page_fans_gender_age = Aggregated demographic data about the people who like your Page based on the age and gender information they provide in their user profiles.
    page_fans_gender_age = graph.get_object(id='974146599436745', fields='insights.metric(page_fans_gender_age)')
    
    #print (json.dumps(page_fans_locale["insights"]["data"][0]['values'][1], indent=4))
def put_zeros_in_empty_cells(file_name):
            # load demo.xlsx 
        wb=load_workbook(file_name)
            # select demo.xlsx
        sheet=wb.active
            # get max row count
        max_row=sheet.max_row
            # get max column count
        max_column=sheet.max_column
        for i in range(1,max_row+1):
        # iterate over all columns
            for j in range(1,max_column+1):
                # get particular cell value    
                if sheet.cell(row=i,column=j).value==None:
                    sheet.cell(row=i,column=j).value = 0
        wb.save(file_name)

def create_age_gender(): #working 100%

    def get_age():
        page_age_gender_day = {}
        #num_of_days = len(page_fans_gender_age["insights"]["data"][0]['values']) #how many days' data i can get
        #for i in range(num_of_days):
        age_gender_time=page_fans_gender_age["insights"]["data"][0]['values'][1]['end_time']
        age_gender_value = page_fans_gender_age["insights"]["data"][0]['values'][1]['value']
            
        temp_dict_age_gender_day = {age_gender_time:age_gender_value}
        page_age_gender_day.update(temp_dict_age_gender_day)

        age_gender_path_direc="excels/lite/"+"lite-Ages-Gender.xlsx"
        #write_age_gender_in_xlsx(page_age_gender_day,age_gender_path_direc)
        #calling functions
            #content
        #print(page_age_gender_day)
        
        if os.path.isfile(age_gender_path_direc):
            update_age_gender_in_xlsx(page_age_gender_day,age_gender_path_direc)
        else:
            write_age_gender_in_xlsx(page_age_gender_day,age_gender_path_direc)  
        
    def write_age_gender_in_xlsx(dictionary, name_xlsx):
        workbook = xlsxwriter.Workbook(name_xlsx)
        worksheet = workbook.add_worksheet()
        #writting first line with date 
        worksheet.write(0, 0,'Date')
        date=list(dictionary.keys())[0]
        date_to_write= date.split("T",1)[0] #old
        date_to_write = datetime.strptime(date_to_write, '%Y-%m-%d').strftime(date_format)
        
        All_Ages = list(dictionary[date].keys())
        row=0
        col=1
        #writting first and second line with gender-age indicators & values of ages-gender
        for i in range(len(All_Ages)):
            
            worksheet.write(row,col,All_Ages[i])
            worksheet.write(row+1,col,dictionary[date][All_Ages[i]])
            col +=1
        workbook.close()
        print(name_xlsx + " is created!")

    def update_age_gender_in_xlsx(dictionary, name_xlsx):
        wb2 = load_workbook(name_xlsx)
        ws = wb2.active
        last_row = ws.max_row
        date=list(dictionary.keys())[0]
        date_to_write= date.split("T",1)[0] #date without the T e.g. 2019-03-05T08:00:00+0000->2019-03-05
        date_to_write = datetime.strptime(date_to_write, '%Y-%m-%d').strftime(date_format)
        ws.cell(last_row+1,1).value = date_to_write

        All_Ages = list(dictionary[date].keys())
        col=2
        for item in range(len(All_Ages)):
            found=False
            minpos=item+col
            while(found==False):
                if(ws.cell(1,minpos).value == All_Ages[item]):
                    posisition_found=minpos
                    found=True
                    #print('Found: {a} at {b}'.format(a=All_Ages[item], b=posisition_found))
                    ws.cell(last_row+1,posisition_found).value = dictionary[date][All_Ages[item]]
                minpos=minpos+1
        wb2.save(name_xlsx)
        #put_zeros_in_empty_cells(name_xlsx)
        put_zeros_in_empty_cells(name_xlsx)
        print(name_xlsx + " is updated!")   

    get_age()

def create_locale(): #working 100%

    def get_locale():
        page_locale_day={}
        locale_time = page_fans_locale["insights"]["data"][0]['values'][1]['end_time']
        locale_value = page_fans_locale["insights"]["data"][0]['values'][1]['value']
        page_locale_day = {locale_time:locale_value}

        locale_path_direct="excels/lite/"+"lite-Locale.xlsx"

        
        if os.path.isfile(locale_path_direct):
            update_locale_in_xlsx(page_locale_day,locale_path_direct)
        else:
            write_locale_in_xlsx(page_locale_day,locale_path_direct)  
        
    def write_locale_in_xlsx(dictionary, name_xlsx):
        workbook = xlsxwriter.Workbook(name_xlsx)
        worksheet = workbook.add_worksheet()
        #writting first line with date 
        worksheet.write(0, 0,'Date')
        date=list(dictionary.keys())[0]
        date_to_write= date.split("T",1)[0] #old
        date_to_write = datetime.strptime(date_to_write, '%Y-%m-%d').strftime(date_format)
        #date_to_write=date_to_write.strftime("%d/%m/%y")
        worksheet.write(1, 0,date_to_write)
        
        All_Locales = list(dictionary[date].keys())
        row=0
        col=1
        #writting first and second line with gender-age indicators & values of ages-gender
        for i in range(len(All_Locales)):
            
            worksheet.write(row,col,All_Locales[i])
            worksheet.write(row+1,col,dictionary[date][All_Locales[i]])
            col +=1
        workbook.close()
        print(name_xlsx + " is created!")

    def update_locale_in_xlsx(dictionary, name_xlsx): #checks also for new entries in locale
        wb2 = load_workbook(name_xlsx)
        ws = wb2.active
        last_row = ws.max_row
        #write date to 1st column last row
        date=list(dictionary.keys())[0]
        date_to_write= date.split("T",1)[0] #old
        date_to_write = datetime.strptime(date_to_write, '%Y-%m-%d').strftime(date_format)
        
        ws.cell(last_row+1,1).value = date_to_write

        All_locale = list(dictionary[date].keys())
        col=2
        already_locale_pos_xlsx = {} 
        # dict for locale : position in first line of xlsx 
        # to know where to add each new locale
        for pos in range(col,ws.max_column+1):
            temp_dict = {ws.cell(1,pos).value : pos}
            already_locale_pos_xlsx.update(temp_dict)
        #print(already_locale_pos_xlsx.keys())
        
        for i in range(len(All_locale)):
            #if the new locale are already in xlsx
            if  All_locale[i] in list(already_locale_pos_xlsx):
                #print("found"+str(All_locale[i])+" at "+str(already_locale_pos_xlsx[All_locale[i]]))
                
                #last row,position that has been found from dictionary
                #value = value from dictionary with key=date ,sub key=locale 
                ws.cell(last_row+1,
                already_locale_pos_xlsx[All_locale[i]]).value = dictionary[date][All_locale[i]]
            
            else :
                #write new locale name in 1st line with other and at that position last_row its value
                last_col = ws.max_column+1

                #print("not found"+str(All_locale[i])+" at "+ str(last_col))

                ws.cell(1,last_col).value = All_locale[i]
                ws.cell(last_row+1,last_col).value = dictionary[date][All_locale[i]]
                last_col+=1
        wb2.save(name_xlsx)
        #put_zeros_in_empty_cells(name_xlsx)
        put_zeros_in_empty_cells(name_xlsx)
        print(name_xlsx + " is updated!")   

    get_locale()

def create_city_country():
    
    def get_city():
        page_city_day={}
        
        city_time=page_fans_city["insights"]["data"][0]['values'][1]['end_time']
        city_value=page_fans_city["insights"]["data"][0]['values'][1]['value']

        temp_dict_age_gender_day = {city_time:city_value}
        page_city_day.update(temp_dict_age_gender_day)

        city_path_direc="excels/lite/"+"lite-City.xlsx"
        #write_city_in_xlsx(page_age_gender_day,age_gender_path_direc)
        #calling functions
            #content
        
        if os.path.isfile(city_path_direc):
            update_city_or_country_in_xlsx(page_city_day,city_path_direc)
        else:
            write_city_or_country_in_xlsx(page_city_day,city_path_direc)  

    def get_country():
        page_country_day={}
        
        country_time=page_fans_country["insights"]["data"][0]['values'][1]['end_time']
        country_value=page_fans_country["insights"]["data"][0]['values'][1]['value']

        temp_dict_age_gender_day = {country_time:country_value}
        page_country_day.update(temp_dict_age_gender_day)

        country_path_direc="excels/lite/"+"lite-Country.xlsx"
        #write_city_in_xlsx(page_age_gender_day,age_gender_path_direc)
        #calling functions
            #content
        
        if os.path.isfile(country_path_direc):
            update_city_or_country_in_xlsx(page_country_day,country_path_direc)
        else:
            write_city_or_country_in_xlsx(page_country_day,country_path_direc) 

    def write_city_or_country_in_xlsx(dictionary, name_xlsx):
        workbook = xlsxwriter.Workbook(name_xlsx)
        worksheet = workbook.add_worksheet()
        #writting first line with date 
        worksheet.write(0, 0,'Date')
        date=list(dictionary.keys())[0]
        date_to_write= date.split("T",1)[0] #old
        date_to_write = datetime.strptime(date_to_write, '%Y-%m-%d').strftime(date_format)
        worksheet.write(1, 0,date_to_write)

        All_Cities = list(dictionary[date].keys())
        row=0
        col=1
        #writting first and second line with gender-age indicators & values of ages-gender
        for i in range(len(All_Cities)):
            
            worksheet.write(row,col,All_Cities[i])
            worksheet.write(row+1,col,dictionary[date][All_Cities[i]])
            col +=1
        workbook.close()
        print(name_xlsx + " is created!")

    def update_city_or_country_in_xlsx(dictionary, name_xlsx):
        wb2 = load_workbook(name_xlsx)
        ws = wb2.active
        last_row = ws.max_row
        #write date to 1st column last row
        date=list(dictionary.keys())[0]
        date_to_write= date.split("T",1)[0] #old
        date_to_write = datetime.strptime(date_to_write, '%Y-%m-%d').strftime(date_format)
        ws.cell(last_row+1,1).value = date_to_write

        All_cities = list(dictionary[date].keys())
        col=2
        already_city_pos_xlsx = {} 
        # dict for cities : position in first line of xlsx 
        # to know where to add each new city
        for pos in range(col,ws.max_column+1):
            temp_dict = {ws.cell(1,pos).value : pos}
            already_city_pos_xlsx.update(temp_dict)
        #print(already_city_pos_xlsx.keys())

        for i in range(len(All_cities)):
            #if the new locale are already in xlsx
            if  All_cities[i] in list(already_city_pos_xlsx):
                #print("found"+str(All_locale[i])+" at "+str(already_locale_pos_xlsx[All_locale[i]]))
                
                #last row,position that has been found from dictionary
                #value = value from dictionary with key=date ,sub key=locale 
                ws.cell(last_row+1,
                already_city_pos_xlsx[All_cities[i]]).value = dictionary[date][All_cities[i]]
            
            else :
                #write new locale name in 1st line with other and at that position last_row its value
                last_col = ws.max_column+1

                #print("not found"+str(All_locale[i])+" at "+ str(last_col))

                ws.cell(1,last_col).value = All_cities[i]
                ws.cell(last_row+1,last_col).value = dictionary[date][All_cities[i]]
                last_col+=1
        wb2.save(name_xlsx)
        #put_zeros_in_empty_cells(name_xlsx)
        print(name_xlsx + " is updated!")
        put_zeros_in_empty_cells(name_xlsx)

    get_city()
    get_country()


def all_functions():
    create_age_gender()
    create_locale()
    create_city_country()

all_functions()