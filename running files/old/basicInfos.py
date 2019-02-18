import os,sys
import json
import facebook
import xlsxwriter
import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd

if __name__ == '__main__':
    token = "EAANXc609TdkBAO3HmSoswBZCTIbmZBMOcdzvLa8c97fdDZBzCjZCL2vAhJYPhyKt5sURY5VlozyHOZABZB6lxrPU5Bb8jM0PLFHh0xCj376nqu6EQZA6PoGbnI1cKyGYiOtrNNyLUebm55GGjNGI5VL6Tj1R9IstsIUSQHBbW7WVP7ZBUbZAn4occ"
    

    graph = facebook.GraphAPI(access_token=token, version = 3.0)
    fan_count = graph.get_object(id='974146599436745', fields='fan_count')
    new_like_count = graph.get_object(id='974146599436745', fields='new_like_count')
    page_impressions = graph.get_object(id='974146599436745', fields='insights.metric(page_impressions).period(day)')
    page_engaged_users = graph.get_object(id='974146599436745', fields='insights.metric(page_engaged_users).period(day)')
    page_views_total = graph.get_object(id='974146599436745', fields='insights.metric(page_views_total).period(day)')
    #print (json.dumps(page_engaged_users, indent=4))
    #print (json.dumps(page_impressions, indent=4))

def update_xlsx(dictionary, name_xlsx,title1, title2, title3, title4 ):
    wb2 = load_workbook(name_xlsx)
    ws = wb2.active
    
    row = ws.max_row
    
    row=row+1
    col = 1
    kkey =list(dictionary.keys())
    
    if (ws.cell(ws.max_row, col).value != kkey[0]) :
        for key in dictionary:
            ws.cell(row, col).value = key
            ws.cell(row, col + 1).value = dictionary[key][0]
            ws.cell(row, col + 2).value = dictionary[key][1]
            
            if title4 != "none":
                ws.cell(row, col + 3).value = dictionary[key][2]
            
    wb2.save(name_xlsx)
    print(name_xlsx +"  is "+ "updated!")


#Write Data in FILE FORMAT: DATE | TITLE1 | TITLE 2
def write_in_Xlsx_3Row( dictionary, name_xlsx, title1, title2, title3, title4 ):
    workbook = xlsxwriter.Workbook(name_xlsx)
    worksheet = workbook.add_worksheet()
    worksheet.write(0, 0,title1)
    worksheet.write(0, 0+1,title2)
    worksheet.write(0, 0+2,title3)
    if title4 != "none":
        worksheet.write(0, 0+3,title4)
    row = 0
    col = 0

    order=sorted(dictionary.keys())
    for key in order:
        row += 1 
        worksheet.write(row, col,     key)
        worksheet.write(row, col + 1, dictionary[key][0])
        worksheet.write(row, col + 2, dictionary[key][1])
        if title4 != "none":
            worksheet.write(row, col + 3, dictionary[key][2])
    workbook.close()
    print(name_xlsx + "is created!")

def create_FanCount_NewLikes():
    fan_count_new_likes_Dict = {}
    temp_date=datetime.datetime.today().strftime("%d/%m/%Y")
    temp_dict={temp_date:
    [fan_count["fan_count"],new_like_count["new_like_count"]]}
    fan_count_new_likes_Dict.update(temp_dict)

    path="/Users/giorgosandreou/Documents/Visual Studio/Facebook data mining/excels/"+"Fan-NewLikes.xlsx"
    
    print(os.path.isfile(path))
    if os.path.isfile(path):
        update_xlsx(temp_dict,path,"Date","Fans Count","New Likes","none")
    else:
        write_in_Xlsx_3Row(fan_count_new_likes_Dict,path, "Date","Fans Count","New Likes","none")
    


def create_Views_Impressions_Engaged():
    end_time2 = page_views_total["insights"]["data"][0]['values'][0]['end_time']
    end_time1 = page_views_total["insights"]["data"][0]['values'][1]['end_time']
    end_time2 = end_time2.split("T",1)[0] #old
    end_time1 = end_time1.split("T",1)[0] #latest
    
    end_time1= datetime.datetime.strptime(end_time1, "%Y-%m-%d").strftime("%d/%m/%Y")
    end_time2= datetime.datetime.strptime(end_time2, "%Y-%m-%d").strftime("%d/%m/%Y")

    page_views_total_value2 = page_views_total["insights"]["data"][0]['values'][0]['value']
    page_views_total_value1 = page_views_total["insights"]["data"][0]['values'][1]['value'] #latest

    page_impressions_value2 = page_impressions["insights"]["data"][0]['values'][0]['value']
    page_impressions_value1 = page_impressions["insights"]["data"][0]['values'][1]['value'] #latest
    
    page_engaged_users_value2 = page_engaged_users["insights"]["data"][0]['values'][0]['value']
    page_engaged_users_value1 = page_engaged_users["insights"]["data"][0]['values'][1]['value'] #latest
    
    page_impressions_views_engaged_Dict = {}
    temp_dict1={end_time1:[page_views_total_value1,page_impressions_value1,page_engaged_users_value1]}
    temp_dict2={end_time2:[page_views_total_value2,page_impressions_value2,page_engaged_users_value2]}
    
    page_impressions_views_engaged_Dict.update(temp_dict2)
    page_impressions_views_engaged_Dict.update(temp_dict1)
    path_direc="/Users/giorgosandreou/Documents/Visual Studio/Facebook data mining/excels/"+"Impression-View-Engaged.xlsx"

    print(os.path.isfile(path_direc))
    if os.path.isfile(path_direc):
        update_xlsx(temp_dict1,path_direc,"Date","Total Views","Impressions","Engaged Users")
    else:
        write_in_Xlsx_3Row(page_impressions_views_engaged_Dict,path_direc,"Date","Total Views","Impressions","Engaged Users")
    


create_FanCount_NewLikes()
create_Views_Impressions_Engaged()