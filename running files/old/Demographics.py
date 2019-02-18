import os,sys
import json
import facebook
import xlsxwriter
import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd


if __name__ == '__main__':
    token = "EAANXc609TdkBAO3HmSoswBZCTIbmZBMOcdzvLa8c97fdDZBzCjZCL2vAhJYPhyKt5sURY5VlozyHOZABZB6lxrPU5Bb8jM0PLFHh0xCj376nqu6EQZA6PoGbnI1cKyGYiOtrNNyLUebm55GGjNGI5VL6Tj1R9IstsIUSQHBbW7WVP7ZBUbZAn4occ"
    

    graph = facebook.GraphAPI(access_token=token, version = 3.0)
    #profile = graph.get_object('974146599436745_974147879436617',fields='get_connections')
    likes = graph.get_object(id='974146599436745_974530109398394', fields='shares,likes.summary(true),comments.summary(true)')

    #age stuff
    page_content_activity_by_age_gender_unique = graph.get_object(id='974146599436745', fields='insights.metric(page_content_activity_by_age_gender_unique).date_preset(last_14d)')
    page_impressions_by_age_gender_unique = graph.get_object(id='974146599436745', fields='insights.metric(page_impressions_by_age_gender_unique).date_preset(last_14d)')
    
    #city stuff
    page_impressions_by_city_unique = graph.get_object(id='974146599436745', fields='insights.metric(page_impressions_by_city_unique).date_preset(last_14d)')
    page_content_activity_by_city_unique = graph.get_object(id='974146599436745', fields='insights.metric(page_content_activity_by_city_unique).date_preset(last_14d)')

    #country stuff
    page_content_activity_by_country_unique = graph.get_object(id='974146599436745', fields='insights.metric(page_content_activity_by_country_unique).date_preset(last_14d)')
    page_impressions_by_country_unique = graph.get_object(id='974146599436745', fields='insights.metric(page_impressions_by_country_unique).date_preset(last_14d)')
    
    #language stuff
    page_content_activity_by_locale_unique = graph.get_object(id='974146599436745', fields='insights.metric(page_content_activity_by_locale_unique).date_preset(last_14d)')
    page_impressions_by_locale_unique = graph.get_object(id='974146599436745', fields='insights.metric(page_impressions_by_locale_unique).date_preset(last_14d)')
    
    page_content_activity_by_action_type_unique= graph.get_object(id='974146599436745', fields='insights.metric(page_content_activity_by_action_type_unique).date_preset(last_14d)')
    
    #page_content_activity_by_age_gender_uniquevalue2 = page_content_activity_by_age_gender_unique["insights"]["data"][1]['values'][4]['end_time']
    
    #cities = list(page_impressions_by_city_unique["insights"]['data'][0]['values'][0]['value'].keys())
    #[x.encode('utf-8') for x in cities]
    #print(cities)
    #print(page_impressions_by_age_gender_unique["insights"]["data"][0])
   
    #print (json.dumps(page_impressions_by_city_unique, indent=4))
    #print(page_content_activity_by_city_unique["insights"]["data"][0]['values'][4]['end_time'])
   
    '''
    print (json.dumps(page_impressions_by_city_unique, indent=4))
    print (json.dumps(page_impressions_by_age_gender_unique, indent=4))
    print(page_impressions_by_age_gender_unique["insights"]["data"][0]['values'][10]['value'])
    '''
#last_14d gives 8 results

def create_age():
    page_content_activity_by_age_gender_unique_4days = {}
    page_impressions_by_age_gender_unique_4days = {}
    
    for i in range(4):
        #print(page_content_activity_by_age_gender_unique["insights"]["data"][0]['values'][i]['end_time'])
        content_temp_end_time=page_content_activity_by_age_gender_unique["insights"]["data"][0]['values'][i]['end_time']
        content_temp_value = page_content_activity_by_age_gender_unique["insights"]["data"][0]['values'][i]['value']
        
        impressions_temp_end_time=page_impressions_by_age_gender_unique["insights"]["data"][0]['values'][i]['end_time']
        impressions_temp_end_value=page_impressions_by_age_gender_unique["insights"]["data"][0]['values'][i]['value']
        
        content_temp_dict_day={content_temp_end_time:content_temp_value}
        impressions_temp_dict_day={impressions_temp_end_time:impressions_temp_end_value}
        
        page_content_activity_by_age_gender_unique_4days.update(content_temp_dict_day)
        page_impressions_by_age_gender_unique_4days.update(impressions_temp_dict_day)


    content_orderDates=sorted(page_content_activity_by_age_gender_unique_4days.keys())
    impressions_orderDates=sorted(page_impressions_by_age_gender_unique_4days.keys())

    content_path_direc="excels/"+"Ages-Content.xlsx"
    impressions_path_direc="excels/"+"Ages-Impressions.xlsx"
    
    #find num of max cells
    max=0
    pos=0
    for i in range(len(content_orderDates)):
        if len(page_content_activity_by_age_gender_unique_4days.get(content_orderDates[i])) > max:
            max = len(page_content_activity_by_age_gender_unique_4days.get(content_orderDates[i]))
                       
            pos+=1
    max_cell_in_Ages=sorted(page_content_activity_by_age_gender_unique_4days.get(content_orderDates[pos]))
    
    #calling functions
        #content
    if os.path.isfile(content_path_direc):
        update_age_in_xlsx(page_content_activity_by_age_gender_unique_4days,content_path_direc,max_cell_in_Ages)
    else:
        write_age_in_xlsx(page_content_activity_by_age_gender_unique_4days,content_path_direc,max_cell_in_Ages)

    #impressions
    if os.path.isfile(impressions_path_direc):
        update_age_in_xlsx(page_impressions_by_age_gender_unique_4days,impressions_path_direc,max_cell_in_Ages)
    else:
        write_age_in_xlsx(page_impressions_by_age_gender_unique_4days,impressions_path_direc,max_cell_in_Ages)
    
def write_age_in_xlsx(dictionary, name_xlsx, ages):
    workbook = xlsxwriter.Workbook(name_xlsx)
    worksheet = workbook.add_worksheet()
    worksheet.write(0, 0,'Date')
    i=0
    for items in ages:
        worksheet.write(0, 1+i,ages[i])
        i=i+1
    row = 0
    col = 0

    orderDates=sorted(dictionary.keys())
    for date in orderDates:
        row += 1
        worksheet.write(row, col,     date)
        singleDateAgesDict= dictionary[date]
        orderAges=sorted(singleDateAgesDict.keys())
        
        for ages in orderAges:
            col +=1
            worksheet.write(row, col,singleDateAgesDict[ages])
        col=0
    workbook.close()
    put_zeros_in_empty_cells(name_xlsx)
    print(name_xlsx + "is created!")

def update_age_in_xlsx(dictionary,name_xlsx,ages):
    wb2 = load_workbook(name_xlsx)
    ws = wb2.active
    row = ws.max_row
    col = 1
    kkey =list(dictionary.keys())
    #check if key exists in xlsx and delete it from parsing dictionary
    
    for key in kkey:
        #print(key)
        for j in range(2,ws.max_row+1):
            #print("value {a}".format(a=ws.cell(row=j,column=1).value))
            if ws.cell(row=j,column=1).value==key:
                
                del dictionary[key]
                
                
    orderDates=sorted(dictionary.keys())
    for date in orderDates:
        row += 1
        
        ws.cell(row, col).value = date
        singleDateAgesDict= dictionary[date]
        orderAges=sorted(singleDateAgesDict.keys())
        for ages in orderAges:
            col +=1
            ws.cell(row, col).value = singleDateAgesDict[ages]
        col=1
    wb2.save(name_xlsx)
    put_zeros_in_empty_cells(name_xlsx)
    print(name_xlsx + " is updated!")
  
def put_zeros_in_empty_cells(file_name):
        # load demo.xlsx 
    wb=load_workbook(file_name)
        # select demo.xlsx
    sheet=wb.active
        # get max row count
    max_row=sheet.max_row
        # get max column count
    max_column=sheet.max_column
    for i in range(1,max_row+1):
    # iterate over all columns
        for j in range(1,max_column+1):
            # get particular cell value    
            if sheet.cell(row=i,column=j).value==None:
                sheet.cell(row=i,column=j).value = 0
    wb.save(file_name)

def call_writing_or_update(dictionary,filepath):
        if os.path.isfile(filepath):
            print(os.path.isfile(filepath))
            update_city_or_country_in_xlsx(dictionary,filepath,None)
        else:
            write_city_or_country_in_xlsx(dictionary,filepath)
        
def create_cities():
    page_impressions_by_city_unique_4days = {}
    page_content_activity_by_city_unique_4days= {}

    for i in range(4):
        content_temp_end_time=page_content_activity_by_city_unique["insights"]["data"][0]['values'][i]['end_time']
        content_temp_value=page_content_activity_by_city_unique["insights"]["data"][0]['values'][i]['value']
        
        for k, v in content_temp_value.items():
            k= k.encode('ascii', 'replace')
            #print(k)
        #print (json.dumps(content_temp_value, indent=4))
        #testdict = {k: v.decode("utf-8", "ignore") for k,v in content_temp_value.items()}
        #print(content_temp_value.items())
        
        impression_temp_end_time=page_impressions_by_city_unique["insights"]["data"][0]['values'][i]['end_time']
        impression_temp_value=page_impressions_by_city_unique["insights"]["data"][0]['values'][i]['value']

        for k, v in impression_temp_value.items():
            k= k.encode('ascii', 'replace')

        content_temp_dict_day={content_temp_end_time:content_temp_value}
        impressions_temp_dict_day={impression_temp_end_time:impression_temp_value}

        page_impressions_by_city_unique_4days.update(impressions_temp_dict_day)
        page_content_activity_by_city_unique_4days.update(content_temp_dict_day)
    

    content_city_path_direc="excels/"+"City-Content.xlsx"
    impression_city_path_direc="excels/"+"City-Impression.xlsx"

    impressions_orderDates=sorted(page_impressions_by_city_unique_4days.keys())

    #calling functions
    call_writing_or_update(page_impressions_by_city_unique_4days,impression_city_path_direc)
    call_writing_or_update(page_content_activity_by_city_unique_4days,content_city_path_direc)

def create_countries_language_impression():
    page_impressions_by_country_8days={}
    page_impressions_by_locale_8days={}

    for i in range(8):
        impression_country_end_time=page_impressions_by_country_unique["insights"]["data"][0]['values'][i]['end_time']
        impression_country_value=page_impressions_by_country_unique["insights"]["data"][0]['values'][i]['value']

        impression_language_end_time=page_impressions_by_locale_unique["insights"]["data"][0]['values'][i]['end_time']
        impression_language_value=page_impressions_by_locale_unique["insights"]["data"][0]['values'][i]['value']

        impression_language_dict_day={impression_language_end_time:impression_language_value}
        impression_country_dict_day={impression_country_end_time:impression_country_value}
        
        page_impressions_by_country_8days.update(impression_country_dict_day)
        page_impressions_by_locale_8days.update(impression_language_dict_day)

    impression_country_path="excels/"+"Country-Impression.xlsx"
    impression_language_path="excels/"+"Language-Impression.xlsx"

    call_writing_or_update(page_impressions_by_country_8days,impression_country_path)
    call_writing_or_update(page_impressions_by_locale_8days,impression_language_path)

def update_city_or_country_in_xlsx(dictionary,name_xlsx,mode):
    print("file:",name_xlsx)
    wb2 = load_workbook(name_xlsx)
    ws = wb2.active
    for k, v in dictionary.items():
            k= k.encode('ascii', 'replace')

    
    maxcol=ws.max_column
    maxrow=ws.max_row

    #print(dictionary.keys())
    all_dates =list(dictionary.keys()) # all dates of dictionary
    #check if date_key exists in xlsx and delete it from parsing into dictionary
    for key in all_dates:
        #print(key)
        for j in range(2,ws.max_row+1):
            
            if ws.cell(row=j,column=1).value==key:
                
                del dictionary[key]
                #print("del{a}".format(a=key))
    ordered_NewDates=sorted(dictionary.keys())
    #print(ordered_NewDates)
    #print(json.dumps(dictionary, indent=4))
    
    updated=False
    if bool(dictionary):
        everyCityAlreadyinXlsx=[]
        for i in range(2,maxcol):
            everyCityAlreadyinXlsx.append(ws.cell(1,i).value)
        #print("cities already in xlsx:")
        #print(json.dumps(everyCityAlreadyinXlsx, indent=4))

        for newDate in ordered_NewDates:
            #print("the date:",newDate)
            ws.cell(ws.max_row+1,column=1).value=newDate #write in the last+1 cell the new date
            #print("row at:",ws.max_row+1,ws.cell(ws.max_row,1).value)
            particularDateCities = dictionary.get(newDate) # the cities of this date
            #print(json.dumps(particularDateCities, indent=4))
            

            for key,value in particularDateCities.items():#add values of cities that already are in xlsx
                if key in everyCityAlreadyinXlsx:
                    for i in range(1,ws.max_column): #search where the city is located, which column in xlsx
                        #print("we are at:",i)
                        if key==ws.cell(1,i).value:
                            #print("found city at:",i)
                            position_column=i
                            ws.cell(ws.max_row,position_column).value=value #write in the found positioned cell the value of the city

                        
                if key not in everyCityAlreadyinXlsx: #add new cities in xlsx and their values
                    #print("not",key.encode('ascii', 'replace'))
                    ws.cell(1,ws.max_column+1).value=key #write in last col the new city
                    ws.cell(ws.max_row,ws.max_column).value=value #write in the found positioned cell the value of the city
                    everyCityAlreadyinXlsx.append(key) #add this city that was written in xlsx 
                    '''
                    for i in range(1,ws.max_column): #search where the city is located, which column in xlsx
                        print("we are at:",i)
                        if key==ws.cell(1,i).value:
                            print("found city at:",i)
                            position_column=i
                            ws.cell(ws.max_row,position_column).value=value #write in the found positioned cell the value of the city
                            #print(newDate,key,value,ws.max_row,position_column)
                    '''
        wb2.save(name_xlsx)
        put_zeros_in_empty_cells(name_xlsx)
        if mode==None:
            updated=True
            print(name_xlsx + " is updated!")
    if(not updated):
        print("nothing to update")
    
def write_city_or_country_in_xlsx(dictionary, name_xlsx):
    workbook = xlsxwriter.Workbook(name_xlsx)
    worksheet = workbook.add_worksheet()
    worksheet.write(0, 0,'Date')
    orderDates=sorted(dictionary.keys())
    firstCitiesInTable=sorted(dictionary[orderDates[0]].keys())
    [x.encode('utf-8') for x in firstCitiesInTable]
    i=0
    for city in firstCitiesInTable:
        #print(city)
        worksheet.write(0, 1+i,firstCitiesInTable[i])
        i=i+1    

    print(name_xlsx + " is created!")
    workbook.close()
    update_city_or_country_in_xlsx(dictionary,name_xlsx,"first")




# page_impressions_by_city_unique & page_content_activity_by_city_unique
create_age()
create_cities()
create_countries_language_impression()

